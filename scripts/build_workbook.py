#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Render vibecoded-app review workbooks in two profiles from the canonical
vibecheck_v1 framework mapping (RFC 0001 §3.4).

Profiles:
  reviewer : technical wording, Pass/Partial/Fail/Not tested/N/A, weights, full gate model.
             Percentages support prioritisation but never determine the verdict.
  founder  : plain-question wording, Pass/Fail/N/A, simplified verdict, % as supporting info,
             "ask your developer/specialist" escalation. For a non-technical app owner.

Both editions share: category, severity, and a "How to verify / evidence to request" column.
Usage:
  python3 build_workbook.py --profile reviewer --lang en --out path.xlsx
  python3 build_workbook.py [--outdir DIR]   # builds all 4 (default: ./ )

Requires openpyxl (see requirements.txt): python3 -m pip install openpyxl
"""
import argparse, os, sys

try:
    import openpyxl
except ImportError:
    sys.exit("build_workbook.py needs openpyxl.\n"
             "  python3 -m pip install openpyxl\n"
             "or: python3 -m pip install -r requirements.txt")

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import controls  # canonical vibecheck_v1 framework mapping
from controls import WEIGHT


# The workbook renders from the canonical vibecheck_v1 framework mapping
# (schema/mappings/vibecheck_v1.json), never from the positional tuples in
# items.py: those are generator-only authoring input, and the mapping is their
# lossless projection.
def _canonical_categories():
    """[(category, [(tuple-shaped item...)])] rebuilt from the canonical mapping,
    preserving workbook order (item_number ascending)."""
    mapping = controls.build_framework_mapping()
    by_number = sorted(mapping["entries"], key=lambda e: e["item_number"])
    cats = []
    for entry in by_number:
        cat = entry["category"]
        tup = (entry["severity"],
               entry["wording"]["tech_en"], entry["wording"]["tech_et"],
               entry["wording"]["plain_en"], entry["wording"]["plain_et"],
               entry["wording"]["test_en"], entry["wording"]["test_et"])
        if cats and cats[-1]["number"] == cat["number"]:
            cats[-1]["items"].append(tup)
        else:
            cats.append({"number": cat["number"], "en": cat["en"], "et": cat["et"],
                         "items": [tup]})
    return cats

CATEGORIES = _canonical_categories()


def _canonical_verification():
    """item_number -> (codes, tools) from the canonical mapping."""
    mapping = controls.build_framework_mapping()
    return {e["item_number"]: (e["verification"]["codes"], e["verification"]["tools"])
            for e in mapping["entries"]}

VERIFICATION = _canonical_verification()

STR = {
 "en": {
  "col_cat":"Category","col_no":"#","col_item":"Control","col_q":"Question","col_sev":"Priority",
  "col_weight":"Weight","col_status":"Status","col_notes":"Notes / evidence / N-A reason","col_how":"How to verify / evidence to request","col_vtype":"Verification type","col_tools":"Tool examples",
  "vt":{"AUTO":"Automated","AI":"AI review","MAN":"Guided manual","E2E":"Automated E2E","SPEC":"Specialist"},
  "guide_tab":"Verification guide",
  "sev":{"Critical":"Critical","High":"High","Medium":"Medium","Low":"Low","Triage":"Screening"},
  "pass":"Pass","partial":"Partial","fail":"Fail","nottested":"Not tested","na":"N/A","acc":"Accepted risk",
  "ans":"Answered","spec":"Needs specialist",
  "review_tab":"Review","summary_tab":"Summary",
  "meta":["Application","Reviewer","Date","Platform / tool","Repo / URL"],
  "verdict":"VERDICT",
  # reviewer verdicts
  "v_notrev":"NOT REVIEWED","v_incomplete":"INCOMPLETE REVIEW","v_block":"BLOCK",
  "v_block_high":"BLOCK - RISK ACCEPTANCE REQUIRED","v_fix":"FIX BEFORE RELEASE",
  "v_rc":"REVIEW COMPLETE - NO OPEN FAIL/PARTIAL",
  # founder verdicts
  "f_notrev":"NOT REVIEWED","f_incomplete":"REVIEW INCOMPLETE","f_donot":"DO NOT LAUNCH",
  "f_fix":"FIX BEFORE LAUNCH","f_ready":"REVIEW COMPLETE - NO OPEN FAILURES",
  "m_coverage":"Review coverage (of applicable)","m_passrate":"Control pass-rate (of verified)",
  "m_applicable":"Applicable controls","m_verified":"Controls verified","m_blocking":"Blocking findings (Critical+High fails)",
  "m_chunrev":"Critical/High not reviewed","m_passev":"Critical/High Pass without evidence","m_narat":"N/A or Accepted risk without a reason","m_acc":"Accepted risks (visible, excluded from pass-rate)","m_critacc":"Critical marked Accepted (not allowed - blocks)","m_triage":"Screening questions unanswered",
  "f_answered":"Questions answered (of applicable)","f_passed":"Checks passed (of answered)",
  "f_chblank":"Critical/High still blank","f_critfail":"Critical failures (do not launch)","f_highfail":"High failures (fix first)",
  "sec_scores":"Per-category coverage & pass-rate","cat":"Category","coverage":"Coverage","passrate":"Pass-rate",
  "how_title":"How to use",
  "title_rev":"Technical Pre-Release Review - Vibecoded Application",
  "title_fnd":"Vibe-Coded App Pre-Launch Quality & Risk Checklist",
  "disc_rev":("Risk-oriented PRE-RELEASE SCREEN, not proof the app is secure, GDPR- or EU-AI-Act-compliant, "
              "or ready to ship. Automated results are evidence/findings, not sign-off. A human decides each Pass. "
              "Use the verdict as a gate; use the percentages for coverage and prioritisation only."),
  "disc_fnd":("This checklist helps non-technical app owners find common high-impact problems before launch. "
              "It does NOT replace professional security, legal or compliance review for sensitive, regulated or "
              "high-scale apps. When a row says 'ask your developer/specialist', get evidence - do not self-certify."),
  "how_rev":[
    "1. Fill metadata, then Status + Notes for each control on the Review tab.",
    "2. Status: Pass / Partial / Fail / Not tested / N/A. Screening rows: Answered / Needs specialist.",
    "3. N/A on a Critical or High control REQUIRES a reason in the Notes column.",
    "4. Weights (Critical 5 / High 3 / Medium 2 / Low 1) drive pass-rate only; screening rows are unscored.",
    "5. Gates: unreviewed Critical/High, unsupported Critical/High Pass, missing N/A or Accepted-risk reason, open screening, or coverage < 100% => INCOMPLETE; any Critical Fail or Critical Accepted => BLOCK; any High Fail => BLOCK/RISK ACCEPTANCE; any other Fail/Partial => FIX; otherwise review complete with no open Fail/Partial.",
    "5b. Accepted risk = a conscious decision, not a fix: write WHO accepted, WHY, and a review-by date in Notes. It counts as reviewed but never as a Pass, and stays visible in the counters. Critical items cannot be accepted - fix or escalate. If you think a severity is overstated, say so in the acceptance reason rather than re-rating the item.",
    "6. A high pass-rate never overrides a failed gate. Categories 3, 14, 15 are product-readiness, not security.",
  ],
  "how_fnd":[
    "1. Fill in the details, then answer each question with Pass / Fail / N/A on the Review tab.",
    "2. Use the 'How to verify' column to actually test each one. Blank means NOT checked - never 'probably fine'.",
    "3. Use N/A only when the feature or risk genuinely does not exist, and write why in Notes.",
    "4. Verdict: incomplete coverage or an unsupported Critical/High Pass => REVIEW INCOMPLETE; any Critical failure or accepted Critical => DO NOT LAUNCH; any other failure => FIX BEFORE LAUNCH; otherwise the review is complete with no open failures.",
    "4b. Accepted risk: only after talking to your developer/specialist. Write who accepted it, why, and when to revisit. Critical problems cannot be accepted - they must be fixed.",
    "5. The percentage is supporting information only - it does not decide the verdict.",
    "6. For items marked 'ask your developer/specialist' (backups, key rotation, GDPR, AI Act, payments), get real evidence.",
  ],
  "legend":"Edit only the yellow cells (details + Status + Notes). Everything else is computed.",
 },
 "et": {
  "col_cat":"Kategooria","col_no":"#","col_item":"Kontroll","col_q":"Kusimus","col_sev":"Prioriteet",
  "col_weight":"Kaal","col_status":"Staatus","col_notes":"Margkmed / tondid / 'ei kohaldu' pohjus","col_how":"Kuidas kontrollida / mis tondit kusida","col_vtype":"Kontrolli tuup","col_tools":"Tooriistade naited",
  "vt":{"AUTO":"Automaatne","AI":"AI ulevaatus","MAN":"Juhendatud kasitsi","E2E":"Automaatne E2E","SPEC":"Spetsialist"},
  "guide_tab":"Kontrolli juhend",
  "sev":{"Critical":"Kriitiline","High":"Korge","Medium":"Keskmine","Low":"Madal","Triage":"Soelumine"},
  "pass":"Korras","partial":"Osaline","fail":"Puudulik","nottested":"Testimata","na":"Ei kohaldu","acc":"Aktsepteeritud risk",
  "ans":"Vastatud","spec":"Vajab spetsialisti",
  "review_tab":"Ulevaatus","summary_tab":"Kokkuvote",
  "meta":["Rakendus","Ulevaataja","Kuupaev","Platvorm / tooriist","Repo / URL"],
  "verdict":"OTSUS",
  "v_notrev":"ULE VAATAMATA","v_incomplete":"ULEVAATUS POOLELI","v_block":"BLOKEERI",
  "v_block_high":"BLOKEERI - NOUAB RISKI AKTSEPTEERIMIST","v_fix":"PARANDA ENNE AVALDAMIST",
  "v_rc":"ULEVAATUS VALMIS - PUUDULIKKE/OSALISI POLE",
  "f_notrev":"ULE VAATAMATA","f_incomplete":"ULEVAATUS POOLELI","f_donot":"ARA AVALDA",
  "f_fix":"PARANDA ENNE AVALDAMIST","f_ready":"ULEVAATUS VALMIS - PUUDULIKKE POLE",
  "m_coverage":"Ulevaatuse kaetus (kohaldatavatest)","m_passrate":"Labimismaar (kontrollitutest)",
  "m_applicable":"Kohaldatavaid kontrolle","m_verified":"Kontrollitud","m_blocking":"Blokeerivaid leide (Krit+Korge)",
  "m_chunrev":"Kriitilisi/Korgeid ule vaatamata","m_passev":"Kriitiline/Korge Korras ilma toendita","m_narat":"'Ei kohaldu' voi aktsepteeritud risk ilma pohjuseta","m_acc":"Aktsepteeritud riske (nahtav, labimismaara ei loeta)","m_critacc":"Kriitiline aktsepteeritud (pole lubatud - blokeerib)","m_triage":"Soelumiskusimusi vastamata",
  "f_answered":"Kusimusi vastatud (kohaldatavatest)","f_passed":"Kontrolle labitud (vastatutest)",
  "f_chblank":"Kriitilisi/Korgeid veel tuhjad","f_critfail":"Kriitilisi puudujaake (ara avalda)","f_highfail":"Korgeid puudujaake (paranda enne)",
  "sec_scores":"Kategooriate kaetus ja labimismaar","cat":"Kategooria","coverage":"Kaetus","passrate":"Labimismaar",
  "how_title":"Kuidas kasutada",
  "title_rev":"Tehniline valjalaske-eelne ulevaatus - vibe-koodiga rakendus",
  "title_fnd":"Vibe-koodiga rakenduse avaldamiseelne kvaliteedi- ja riskikontroll",
  "disc_rev":("Riskipohine VALJALASKE-EELNE SKANN, mitte tond, et rakendus on turvaline, GDPR- voi EL AI-maaruse-vastav "
              "voi valmis. Automaatsed tulemused on tondid/leiud, mitte kinnitus. Iga 'Korras' otsustab inimene. "
              "Kasuta otsust varavana; protsente ainult kaetuse ja prioriseerimise jaoks."),
  "disc_fnd":("See kontrollnimekiri aitab mittetehnilisel omanikul leida levinud suure mojuga probleeme enne avaldamist. "
              "See EI asenda professionaalset turva-, oigus- voi vastavusulevaatust tundlike, reguleeritud voi suuremahuliste "
              "rakenduste puhul. Kui rida utleb 'kusi arendajalt/spetsialistilt', hangi tond - ara ise kinnita."),
  "how_rev":[
    "1. Taida metaandmed, siis iga kontrolli Staatus + Margkmed 'Ulevaatus' lehel.",
    "2. Staatus: Korras / Osaline / Puudulik / Testimata / Ei kohaldu. Soelumisread: Vastatud / Vajab spetsialisti.",
    "3. 'Ei kohaldu' Kriitilisel voi Korgel kontrollil NOUAB pohjust Margkmete veerus.",
    "4. Kaalud (Krit 5 / Korge 3 / Keskmine 2 / Madal 1) mojutavad ainult labimismaara; soelumisread on skoorimata.",
    "5. Varavad: ule vaatamata voi toendita Korras Krit/Korge, puuduv pohjus, avatud soelumine voi kaetus < 100% => POOLELI; iga Krit Puudulik voi Krit aktsepteeritud => BLOKEERI; iga Korge Puudulik => BLOKEERI/RISKI AKTSEPT; muu Puudulik/Osaline => PARANDA; muidu ulevaatus valmis.",
    "5b. Aktsepteeritud risk = teadlik otsus, mitte parandus: kirjuta Margkmetesse KES, MIKS ja ulevaatuse kuupaev. Loeb ulevaadatuks, mitte kunagi Korras'eks, ja jaab loenduritesse nahtavaks. Kriitilisi ei saa aktsepteerida - paranda voi eskaleeri. Kui raskusaste tundub ulepaisutatud, pohjenda seda aktsepteerimise pohjuses, mitte umberhindamisega.",
    "6. Korge labimismaar ei tuhista kunagi labikukkunud varavat. Kategooriad 3, 14, 15 on tootevalmidus, mitte turve.",
  ],
  "how_fnd":[
    "1. Taida andmed, siis vasta igale kusimusele Korras / Puudulik / Ei kohaldu 'Ulevaatus' lehel.",
    "2. Kasuta 'Kuidas kontrollida' veergu, et igauht pariselt testida. Tuhi tahendab EI KONTROLLITUD - mitte 'kull sobib'.",
    "3. 'Ei kohaldu' ainult kui funktsiooni voi riski pariselt pole, ja kirjuta miks Margkmetesse.",
    "4. Otsus: puudulik kaetus voi toendita Kriitiline/Korge Korras => POOLELI; iga Kriitiline puudulik voi aktsepteeritud => ARA AVALDA; iga muu puudulik => PARANDA ENNE; muidu ulevaatus valmis ja avatud puudujaake pole.",
    "4b. Aktsepteeritud risk: ainult parast arendaja/spetsialistiga raakimist. Kirjuta, kes aktsepteeris, miks ja millal ule vaadata. Kriitilisi probleeme ei saa aktsepteerida - need tuleb parandada.",
    "5. Protsent on ainult tugiinfo - see ei otsusta tulemust.",
    "6. Ridade puhul 'kusi arendajalt/spetsialistilt' (varukoopiad, votmed, GDPR, AI-maarus, maksed) hangi paris tond.",
  ],
  "legend":"Muuda ainult kollaseid lahtreid (andmed + Staatus + Margkmed). Ulejaanu arvutatakse.",
 },
}

NAVY="1F2A44"; HDR="2E3B55"; YELLOW="FFF2CC"; REF="EEF2F7"
GREEN="C6EFCE"; GREEN_T="006100"; RED="FFC7CE"; RED_T="9C0006"
AMBER="FFEB9C"; GREY="D9D9D9"; GREY_T="808080"
thin=Side(style="thin",color="B7BEC9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

def hcell(c):
    c.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
    c.fill=PatternFill("solid",fgColor=HDR)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border

def build(profile, lang, path):
    t=STR[lang]; sv=t["sev"]; founder=(profile=="founder")
    P,PT,F,NT,NA,ACC=t["pass"],t["partial"],t["fail"],t["nottested"],t["na"],t["acc"]
    ANS,SPEC=t["ans"],t["spec"]; crit,high,scr=sv["Critical"],sv["High"],sv["Triage"]
    item_idx = 3 if founder else 1   # plain question vs technical wording
    test_idx = 6 if lang=="et" else 5

    wb=openpyxl.Workbook(); rv=wb.active; rv.title=t["review_tab"]
    # cols: A cat B # C control D priority E weight F status G notes H how-to-verify
    itemhdr = t["col_q"] if founder else t["col_item"]
    headers=[t["col_cat"],t["col_no"],itemhdr,t["col_sev"],t["col_weight"],t["col_status"],t["col_notes"],t["col_how"],t["col_vtype"],t["col_tools"]]
    for i,h in enumerate(headers,1): hcell(rv.cell(row=1,column=i,value=h))
    rv.freeze_panes="A2"

    r=2; n=0; fdr=2; scr_rows=[]
    for cat in CATEGORIES:
        cname=cat["en"] if lang=="en" else cat["et"]
        for it in cat["items"]:
            sev=it[0]; n+=1
            tech_en,tech_et,plain_en,plain_et,test_en,test_et=it[1],it[2],it[3],it[4],it[5],it[6]
            control = (plain_en if lang=="en" else plain_et) if founder else (tech_en if lang=="en" else tech_et)
            howto = test_en if lang=="en" else test_et
            rv.cell(row=r,column=1,value=cname)
            rv.cell(row=r,column=2,value=n)
            rv.cell(row=r,column=3,value=control)
            rv.cell(row=r,column=4,value=sv[sev])
            rv.cell(row=r,column=5,value=WEIGHT[sev])
            rv.cell(row=r,column=8,value=howto)
            vcodes,vtools=VERIFICATION[n]
            rv.cell(row=r,column=9,value=" + ".join(t["vt"][c] for c in vcodes))
            rv.cell(row=r,column=10,value=vtools)
            if sev=="Triage": scr_rows.append(r)
            for c in range(1,11):
                cell=rv.cell(row=r,column=c); cell.border=border
                cell.font=Font(name="Arial",size=10,bold=(sev=="Critical" and c==4),
                               color=(RED_T if sev=="Critical" and c==4 else "000000"),
                               italic=(c in (8,10)))
                cell.alignment=Alignment(horizontal=("center" if c in (2,4,5) else "left"),
                                         vertical="top",wrap_text=(c in (1,3,7,8,10)))
                if c in (6,7): cell.fill=PatternFill("solid",fgColor=YELLOW)
                if c in (8,9,10): cell.fill=PatternFill("solid",fgColor=REF)
            r+=1
    ldr=r-1

    if founder:
        dv=DataValidation(type="list",formula1=f'"{P},{F},{NA},{ACC}"',allow_blank=True)
    else:
        dv=DataValidation(type="list",formula1=f'"{P},{PT},{F},{NT},{NA},{ACC}"',allow_blank=True)
    dv_t=DataValidation(type="list",formula1=f'"{ANS},{SPEC}"',allow_blank=True)
    rv.add_data_validation(dv); rv.add_data_validation(dv_t)
    for row in range(fdr,ldr+1):
        (dv_t if row in scr_rows else dv).add(f"F{row}")

    Fr=f"F{fdr}:F{ldr}"
    for val,fill,ft in [(P,GREEN,GREEN_T),(F,RED,RED_T),(NA,GREY,GREY_T),(ACC,"E1D5E7","5B3A7E")]:
        rv.conditional_formatting.add(Fr,CellIsRule(operator="equal",formula=[f'"{val}"'],
            fill=PatternFill("solid",fgColor=fill),font=Font(color=ft,bold=(val!=NA))))
    if not founder:
        rv.conditional_formatting.add(Fr,CellIsRule(operator="equal",formula=[f'"{PT}"'],fill=PatternFill("solid",fgColor=AMBER)))
        rv.conditional_formatting.add(Fr,CellIsRule(operator="equal",formula=[f'"{NT}"'],fill=PatternFill("solid",fgColor="FFD9B3")))

    widths={"A":24,"B":4,"C":60,"D":11,"E":7,"F":14,"G":28,"H":46,"I":18,"J":30}
    for col,w in widths.items(): rv.column_dimensions[col].width=w
    if founder: rv.column_dimensions["E"].hidden=True  # weights not shown to owners

    RT=t["review_tab"]
    A=f"'{RT}'!$A${fdr}:$A${ldr}"; D=f"'{RT}'!$D${fdr}:$D${ldr}"; E=f"'{RT}'!$E${fdr}:$E${ldr}"
    Fc=f"'{RT}'!$F${fdr}:$F${ldr}"; G=f"'{RT}'!$G${fdr}:$G${ldr}"

    # shared count formulas (scoreable = weight>0)
    nPass=f'COUNTIFS({Fc},"{P}",{E},">0")'; nFail=f'COUNTIFS({Fc},"{F}",{E},">0")'
    nPartial=f'COUNTIFS({Fc},"{PT}",{E},">0")'
    nNA=f'COUNTIFS({Fc},"{NA}",{E},">0")'; scoreTot=f'COUNTIF({E},">0")'
    critFail=f'COUNTIFS({D},"{crit}",{Fc},"{F}")'; highFail=f'COUNTIFS({D},"{high}",{Fc},"{F}")'
    scrTot=f'COUNTIF({D},"{scr}")'
    scrOpen=f'({scrTot}-COUNTIFS({D},"{scr}",{Fc},"{ANS}")-COUNTIFS({D},"{scr}",{Fc},"{SPEC}"))'
    naNoRat=f'(SUMPRODUCT((({D}="{crit}")+({D}="{high}"))*({Fc}="{NA}")*({G}="")))'
    nAcc=f'COUNTIFS({Fc},"{ACC}",{E},">0")'
    critAcc=f'COUNTIFS({D},"{crit}",{Fc},"{ACC}")'
    accNoRat=f'(SUMPRODUCT(({Fc}="{ACC}")*({G}="")))'
    passNoEv=(f'(SUMPRODUCT((({D}="{crit}")+({D}="{high}"))'
              f'*({Fc}="{P}")*({G}="")))')
    applicable=f'({scoreTot}-{nNA})'

    sm=wb.create_sheet(t["summary_tab"]); wb.move_sheet(t["summary_tab"],-(len(wb.sheetnames)-1))
    sm.sheet_view.showGridLines=False
    title=t["title_fnd"] if founder else t["title_rev"]; disc=t["disc_fnd"] if founder else t["disc_rev"]
    sm.merge_cells("A1:F1"); sm["A1"]=title
    sm["A1"].font=Font(name="Arial",bold=True,size=15,color="FFFFFF"); sm["A1"].fill=PatternFill("solid",fgColor=NAVY)
    sm["A1"].alignment=Alignment(horizontal="left",vertical="center",indent=1); sm.row_dimensions[1].height=28
    sm.merge_cells("A2:F3"); sm["A2"]=disc
    sm["A2"].font=Font(name="Arial",size=9,italic=True,color="7A2E00"); sm["A2"].fill=PatternFill("solid",fgColor="FFF6E6")
    sm["A2"].alignment=Alignment(horizontal="left",vertical="center",indent=1,wrap_text=True)

    r=5
    for label in t["meta"]:
        sm.cell(row=r,column=1,value=label).font=Font(name="Arial",bold=True,size=10)
        sm.cell(row=r,column=1).alignment=Alignment(horizontal="right")
        sm.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
        cc=sm.cell(row=r,column=2); cc.fill=PatternFill("solid",fgColor=YELLOW); cc.border=border
        r+=1

    r+=1
    sm.cell(row=r,column=1,value=t["verdict"]).font=Font(name="Arial",bold=True,size=12)
    sm.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
    v=sm.cell(row=r,column=2)
    if founder:
        chBlank=f'(COUNTIFS({D},"{crit}",{Fc},"")+COUNTIFS({D},"{high}",{Fc},""))'
        answered=f'({nPass}+{nFail}+{nAcc})'
        coverage=f'IF({applicable}=0,0,{answered}/{applicable})'
        anyRev=f'({answered}+{nNA}+({scrTot}-{scrOpen}))'
        v.value=(f'=IF({anyRev}=0,"{t["f_notrev"]}",'
                 f'IF(OR({chBlank}>0,{scrOpen}>0,{naNoRat}>0,{accNoRat}>0,'
                 f'{passNoEv}>0,{coverage}<1),"{t["f_incomplete"]}",'
                 f'IF(({critFail}+{critAcc})>0,"{t["f_donot"]}",'
                 f'IF({nFail}>0,"{t["f_fix"]}","{t["f_ready"]}"))))')
        vmap=[(t["f_notrev"],"808080"),(t["f_incomplete"],"B7791F"),(t["f_donot"],"C00000"),
              (t["f_fix"],"E8A200"),(t["f_ready"],"2E7D32")]
    else:
        verified=f'({nPass}+{nPartial}+{nFail}+{nAcc})'
        coverage=f'IF({applicable}=0,0,{verified}/{applicable})'
        earnedW=f'(SUMIFS({E},{Fc},"{P}")+0.5*SUMIFS({E},{Fc},"{PT}"))'
        applW=f'(SUMIFS({E},{Fc},"{P}")+SUMIFS({E},{Fc},"{PT}")+SUMIFS({E},{Fc},"{F}"))'
        passrate=f'IF({applW}=0,0,{earnedW}/{applW})'
        chUnrev=(f'(COUNTIFS({D},"{crit}",{Fc},"")+COUNTIFS({D},"{crit}",{Fc},"{NT}")'
                 f'+COUNTIFS({D},"{high}",{Fc},"")+COUNTIFS({D},"{high}",{Fc},"{NT}"))')
        anyRev=f'({verified}+{nNA}+({scrTot}-{scrOpen}))'
        v.value=(f'=IF({anyRev}=0,"{t["v_notrev"]}",'
                 f'IF(OR({chUnrev}>0,{passNoEv}>0,{naNoRat}>0,{accNoRat}>0,{scrOpen}>0),"{t["v_incomplete"]}",'
                 f'IF(({critFail}+{critAcc})>0,"{t["v_block"]}",'
                 f'IF({highFail}>0,"{t["v_block_high"]}",'
                 f'IF({coverage}<1,"{t["v_incomplete"]}",'
                 f'IF(({nFail}+{nPartial})>0,"{t["v_fix"]}","{t["v_rc"]}"))))))')
        vmap=[(t["v_notrev"],"808080"),(t["v_incomplete"],"B7791F"),(t["v_block"],"C00000"),
              (t["v_block_high"],"C00000"),(t["v_fix"],"E8A200"),(t["v_rc"],"2E7D32")]
    v.font=Font(name="Arial",bold=True,size=13,color="FFFFFF"); v.alignment=Alignment(horizontal="center",vertical="center")
    v.fill=PatternFill("solid",fgColor="808080"); sm.row_dimensions[r].height=30
    for txt,col in vmap:
        sm.conditional_formatting.add(f"B{r}",CellIsRule(operator="equal",formula=[f'"{txt}"'],
            fill=PatternFill("solid",fgColor=col),font=Font(bold=True,size=13,color="FFFFFF")))

    def metric(rr,label,formula,pct=False,warn=False):
        sm.cell(row=rr,column=1,value=label).font=Font(name="Arial",bold=True,size=10)
        sm.cell(row=rr,column=1).alignment=Alignment(horizontal="right",wrap_text=True)
        c=sm.cell(row=rr,column=2,value=formula); c.font=Font(name="Arial",size=11,bold=True)
        c.alignment=Alignment(horizontal="center"); c.border=border
        if pct: c.number_format="0.0%"
        if warn:
            sm.conditional_formatting.add(f"B{rr}",CellIsRule(operator="greaterThan",formula=["0"],
                fill=PatternFill("solid",fgColor=RED),font=Font(bold=True,color=RED_T)))
    r+=2
    if founder:
        cov=f'IF({applicable}=0,0,{answered}/{applicable})'
        pr=f'IF({answered}=0,0,{nPass}/{answered})'
        metric(r,t["f_answered"],f"={cov}",pct=True); r+=1
        metric(r,t["f_passed"],f"={pr}",pct=True); r+=1
        metric(r,t["f_chblank"],f"={chBlank}",warn=True); r+=1
        metric(r,t["m_passev"],f"={passNoEv}",warn=True); r+=1
        metric(r,t["f_critfail"],f"={critFail}",warn=True); r+=1
        metric(r,t["f_highfail"],f"={highFail}",warn=True); r+=1
        metric(r,t["m_acc"],f"=COUNTIF({Fc},\"{ACC}\")"); r+=1
        metric(r,t["m_critacc"],f"={critAcc}",warn=True); r+=1
        metric(r,t["m_triage"],f"={scrOpen}",warn=True); r+=1
    else:
        metric(r,t["m_coverage"],f"={coverage}",pct=True); r+=1
        metric(r,t["m_passrate"],f"={passrate}",pct=True); r+=1
        metric(r,t["m_applicable"],f"={applicable}"); r+=1
        metric(r,t["m_verified"],f"={verified}"); r+=1
        metric(r,t["m_blocking"],f"=({critFail}+{highFail})",warn=True); r+=1
        metric(r,t["m_chunrev"],f"={chUnrev}",warn=True); r+=1
        metric(r,t["m_passev"],f"={passNoEv}",warn=True); r+=1
        metric(r,t["m_narat"],f"=({naNoRat}+{accNoRat})",warn=True); r+=1
        metric(r,t["m_acc"],f"=COUNTIF({Fc},\"{ACC}\")"); r+=1
        metric(r,t["m_critacc"],f"={critAcc}",warn=True); r+=1
        metric(r,t["m_triage"],f"={scrOpen}",warn=True); r+=1

    # per-category coverage + pass-rate
    r+=1
    sm.cell(row=r,column=1,value=t["sec_scores"]).font=Font(name="Arial",bold=True,size=12,color=NAVY); r+=1
    for i,lab in enumerate([t["cat"],t["coverage"],t["passrate"]],1): hcell(sm.cell(row=r,column=i,value=lab))
    r+=1; cov_first=r
    for cat in CATEGORIES:
        cname=cat["en"] if lang=="en" else cat["et"]
        is_scr=all(x[0]=="Triage" for x in cat["items"])
        sm.cell(row=r,column=1,value=cname).font=Font(name="Arial",size=9)
        sm.cell(row=r,column=1).alignment=Alignment(horizontal="left",wrap_text=True); sm.cell(row=r,column=1).border=border
        if is_scr:
            so=f'({scrTot}-COUNTIFS({D},"{scr}",{Fc},"{ANS}")-COUNTIFS({D},"{scr}",{Fc},"{SPEC}"))'
            cc=sm.cell(row=r,column=2,value=f'=IF({so}=0,"screening done","screening: "&{so}&" open")')
            cc.font=Font(name="Arial",size=9,italic=True); cc.alignment=Alignment(horizontal="center"); cc.border=border
            sm.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
        else:
            tot=f'COUNTIFS({A},$A{r},{E},">0")'; na=f'COUNTIFS({A},$A{r},{Fc},"{NA}",{E},">0")'
            if founder:
                ver=f'(COUNTIFS({A},$A{r},{Fc},"{P}")+COUNTIFS({A},$A{r},{Fc},"{F}")+COUNTIFS({A},$A{r},{Fc},"{ACC}"))'
                ew=f'COUNTIFS({A},$A{r},{Fc},"{P}")'
                aw=f'(COUNTIFS({A},$A{r},{Fc},"{P}")+COUNTIFS({A},$A{r},{Fc},"{F}"))' 
            else:
                ver=f'(COUNTIFS({A},$A{r},{Fc},"{P}")+COUNTIFS({A},$A{r},{Fc},"{PT}")+COUNTIFS({A},$A{r},{Fc},"{F}")+COUNTIFS({A},$A{r},{Fc},"{ACC}"))'
                ew=f'(SUMIFS({E},{A},$A{r},{Fc},"{P}")+0.5*SUMIFS({E},{A},$A{r},{Fc},"{PT}"))'
                aw=f'(SUMIFS({E},{A},$A{r},{Fc},"{P}")+SUMIFS({E},{A},$A{r},{Fc},"{PT}")+SUMIFS({E},{A},$A{r},{Fc},"{F}"))'
            appl=f'({tot}-{na})'
            cvc=sm.cell(row=r,column=2,value=f'=IF({appl}=0,"n/a",{ver}&"/"&{appl})')
            cvc.font=Font(name="Arial",size=9); cvc.alignment=Alignment(horizontal="center"); cvc.border=border
            prc=sm.cell(row=r,column=3,value=f'=IF({aw}=0,"-",{ew}/{aw})')
            prc.number_format="0%"; prc.font=Font(name="Arial",size=9,bold=True)
            prc.alignment=Alignment(horizontal="center"); prc.border=border
        r+=1
    cov_last=r-1
    sm.conditional_formatting.add(f"C{cov_first}:C{cov_last}",CellIsRule(operator="lessThan",formula=["0.7"],fill=PatternFill("solid",fgColor=RED)))
    sm.conditional_formatting.add(f"C{cov_first}:C{cov_last}",CellIsRule(operator="between",formula=["0.7","0.899"],fill=PatternFill("solid",fgColor=AMBER)))
    sm.conditional_formatting.add(f"C{cov_first}:C{cov_last}",CellIsRule(operator="greaterThanOrEqual",formula=["0.9"],fill=PatternFill("solid",fgColor=GREEN)))

    r+=1
    sm.cell(row=r,column=1,value=t["how_title"]).font=Font(name="Arial",bold=True,size=11,color=NAVY); r+=1
    for line in (t["how_fnd"] if founder else t["how_rev"]):
        sm.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        sm.cell(row=r,column=1,value=line).font=Font(name="Arial",size=9,color="333333")
        sm.cell(row=r,column=1).alignment=Alignment(horizontal="left",wrap_text=True); sm.row_dimensions[r].height=24; r+=1
    sm.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
    sm.cell(row=r,column=1,value=t["legend"]).font=Font(name="Arial",size=9,italic=True,color="7A6000")
    sm.cell(row=r,column=1).fill=PatternFill("solid",fgColor=YELLOW)
    sm.cell(row=r,column=1).alignment=Alignment(horizontal="left",indent=1,wrap_text=True)

    sm.column_dimensions["A"].width=40
    for col in ["B","C","D","E","F"]: sm.column_dimensions[col].width=15
    _guide(wb, t, lang)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(path); return n


GUIDE = {
 "en": {
  "t1":"Verification types","t2":"Recommended tool stack",
  "h1":["Type","What it means","Typical tools","Good for","Do NOT assume","Owner action"],
  "r1":[
   ["Automated","A tool checks code, configuration or the running site repeatedly.","Gitleaks, Semgrep Community, OSV-Scanner, Trivy, OWASP ZAP, axe","Secrets, dependencies, headers, obvious code patterns, asset size","A green scan proves the app is safe","Run locally or in CI, review every finding, rerun after changes"],
   ["AI review","An AI coding assistant reads the repository and explains likely issues.","Claude Code, Cursor, GitHub Copilot, ChatGPT/Codex","Locating relevant code, unsafe patterns, duplicated logic, test suggestions","The AI has tested production behaviour or understood every business rule","Ask for file/line evidence and verify important claims yourself"],
   ["Guided manual","A person follows a concrete test using accounts, browser tools or provider dashboards.","Two test accounts, browser DevTools, real inbox/phone, provider test mode","Authorization, persistence, payments, email, uploads, failure handling","A developer statement is equivalent to a successful test","Record result, screenshot and date in the Notes column"],
   ["Automated E2E","A browser robot repeats a small number of core user journeys.","Playwright","Signup, login, create/edit, upload, payment, the primary workflow","Every changing UI detail needs a large test suite","Keep 5-10 stable critical-flow tests, run on each release"],
   ["Specialist","The answer needs operational evidence, business intent or legal/security judgement.","Developer, security reviewer, privacy/legal adviser, platform support","Backup restore, key rotation, GDPR/AI Act, incident response","The checklist can self-certify regulated or high-impact use","Escalate uncertain Critical/High answers before launch"],
  ],
  "h2":["Tool","Purpose","When to use","Coverage","Limitation","Practical setup"],
  "r2":[
   ["Gitleaks + TruffleHog","Free/open-source secret scanning","Every repository and its full history","Known and high-entropy credential patterns","A hit still needs validity/placement review; rotate real leaks","Run locally and in CI over full history"],
   ["Semgrep Community / CodeQL","Static code analysis; CodeQL is free for public GitHub repos","Language-supported codebases","Injection, insecure APIs and framework-specific patterns","No ruleset understands every business or authorization rule","Start with maintained security rules; triage with code context"],
   ["OSV-Scanner / Trivy","Free/open-source dependency and supply-chain scanning","Lockfiles, containers and CI","Known advisories, vulnerable packages and images","Advisory presence is not exploitability; absence is not safety","Scan lockfiles/images in CI and triage reachability"],
   ["OWASP ZAP","Free/open-source dynamic web scanner","A test deployment you are authorized to probe","Headers, exposed routes and common web vulnerabilities","Can change state or create load; does not understand business intent","Use passive/baseline mode first against staging"],
   ["Playwright","Free/open-source browser-flow tests","Critical user journeys","Signup, login, persistence, authorization and regressions","Expected outcomes must be designed by a human","Keep a small stable suite using two test accounts"],
   ["Sentry","Production error and performance monitoring","Any public MVP","Real client/server failures and regressions","Detects failures after they occur; not a pre-launch test","Trigger one test error and confirm the alert arrives"],
   ["axe/WAVE + Lighthouse","Accessibility and frontend quality","Any web interface","Basic accessibility, asset and performance issues","Automated accessibility checks are incomplete","Run before launch and after major UI changes"],
   ["vibecheck","Review orchestration and lightweight static signals","Every review, as a checklist driver","Maps signals and manual work to this checklist","Less capable than dedicated scanners; NO_SIGNAL is not a Pass","Run scripts/vibecheck.sh; confirm WARN and complete MANUAL checks"],
  ],
 },
 "et": {
  "t1":"Kontrolli tuubid","t2":"Soovitatav tooriistakomplekt",
  "h1":["Tuup","Mida tahendab","Tuupilised tooriistad","Sobib","ARA eelda","Omaniku tegevus"],
  "r1":[
   ["Automaatne","Tooriist kontrollib koodi, seadistust voi tootavat saiti korduvalt.","Gitleaks, Semgrep Community, OSV-Scanner, Trivy, OWASP ZAP, axe","Saladused, soltuvused, paised, ilmsed koodimustrid, varade suurus","Roheline skann toestab, et rakendus on turvaline","Kaivita kohapeal voi CI-s, vaata iga leid ule ja korda parast muudatusi"],
   ["AI ulevaatus","AI-abiline loeb hoidlat ja selgitab tenaoseid probleeme.","Claude Code, Cursor, GitHub Copilot, ChatGPT/Codex","Asjakohase koodi leidmine, ohtlikud mustrid, dubleeritud loogika, testisoovitused","AI on testinud tootmiskaitumist voi moistnud iga arireeglit","Kusi faili/rea toendit ja kontrolli olulised vaited ise"],
   ["Juhendatud kasitsi","Inimene labib konkreetse testi kontode, brauseri tooriistade voi tooulaudade abil.","Kaks testkontot, brauseri DevTools, paris postkast/telefon, teenuse testreziim","Volitamine, pusivus, maksed, e-post, uleslaadimised, torgete kaitumine","Arendaja vaide vordub onnestunud testiga","Salvesta tulemus, ekraanipilt ja kuupaev Margkmete veergu"],
   ["Automaatne E2E","Brauserirobot kordab vaikest arvu pohilisi kasutajateid.","Playwright","Registreerimine, login, loomine/muutmine, uleslaadimine, makse, pohivoog","Iga muutuv liidese detail vajab suurt testikomplekti","Hoia 5-10 stabiilset kriitilise voo testi, kaivita igal valjalaskel"],
   ["Spetsialist","Vastus vajab operatiivset toendit, arikavatsust voi oigus-/turvahinnangut.","Arendaja, turvaulevaataja, privaatsus-/oigusnounik, platvormi tugi","Varukoopia taaste, votmete rotatsioon, GDPR/AI-maarus, intsidendid","Kontrollnimekiri saab ise kinnitada reguleeritud voi suure mojuga kasutust","Eskaleeri ebakindlad Krit/Korge vastused enne avaldamist"],
  ],
  "h2":["Tooriist","Otstarve","Millal kasutada","Kaetus","Piirang","Praktiline seadistus"],
  "r2":[
   ["Gitleaks + TruffleHog","Tasuta/avatud lahtekoodiga saladuste skann","Iga hoidla ja kogu selle ajalugu","Tuntud ja suure entroopiaga mandaadimustrid","Leid vajab kehtivuse/asukoha kontrolli; paris leke roteeri","Kaivita kohapeal ja CI-s kogu ajaloo peal"],
   ["Semgrep Community / CodeQL","Staatiline koodianaluus; CodeQL on avalikel GitHubi hoidlail tasuta","Toetatud keeltega koodibaasid","Injektsioon, ohtlikud API-d ja raamistikumustrid","Ukski reeglistik ei moista koiki ari- ega volitusreegleid","Alusta hooldatud turvareeglitest; hinda koodikontekstis"],
   ["OSV-Scanner / Trivy","Tasuta/avatud soltuvus- ja tarneahelaskann","Lukufailid, konteinerid ja CI","Tuntud turvateated, haavatavad paketid ja tommised","Turvateade ei toesta kasutatavust; puudumine ei toesta turvalisust","Skanni lukufaile/tommiseid CI-s ja hinda ulatuvust"],
   ["OWASP ZAP","Tasuta/avatud dunaamiline veebiskanner","Testjuurutus, mida tohib skannida","Paisid, avatud marsruudid ja levinud veebihaavatavused","Voib muuta olekut voi tekitada koormust; arikavatsust ei moista","Alusta passiivse/baseline-reziimiga stagingus"],
   ["Playwright","Tasuta/avatud brauserivoogude testid","Kriitilised kasutajateekonnad","Registreerimine, login, pusivus, volitamine, regressioonid","Oodatavad tulemused peab inimene kavandama","Hoia vaike stabiilne komplekt kahe testkontoga"],
   ["Sentry","Tootmisvigade ja joudluse seire","Iga avalik MVP","Paris kliendi-/serveritorked ja regressioonid","Tuvastab torked parast toimumist; pole avaldamiseelne test","Vallanda uks testviga ja kinnita teate saabumine"],
   ["axe/WAVE + Lighthouse","Ligipaasetavus ja frontendi kvaliteet","Iga veebiliides","Pohiligipaasetavus, varade ja joudluse probleemid","Automaatsed ligipaasetavuse kontrollid on mittetaielikud","Kaivita enne avaldamist ja parast suuremaid UI-muudatusi"],
   ["vibecheck","Ulevaatuse orkestreerimine ja kerged staatilised signaalid","Iga ulevaatus, kontrollnimekirja juhtimiseks","Seob signaalid ja kasitoo selle nimekirjaga","Norgem kui spetsiaalsed skannerid; NO_SIGNAL pole Korras","Kaivita scripts/vibecheck.sh; kinnita WARN ja tee MANUAL kontrollid"],
  ],
 },
}

def _guide(wb, t, lang):
    g=GUIDE[lang]
    ws=wb.create_sheet(t["guide_tab"]); ws.sheet_view.showGridLines=False
    r=1
    for title,hdrs,rows in [(g["t1"],g["h1"],g["r1"]),(g["t2"],g["h2"],g["r2"])]:
        ws.cell(row=r,column=1,value=title).font=Font(name="Arial",bold=True,size=13,color=NAVY); r+=1
        for i,h in enumerate(hdrs,1): hcell(ws.cell(row=r,column=i,value=h))
        r+=1
        for row in rows:
            for i,val in enumerate(row,1):
                c=ws.cell(row=r,column=i,value=val)
                c.font=Font(name="Arial",size=9,bold=(i==1))
                c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
                c.border=border
            ws.row_dimensions[r].height=42; r+=1
        r+=2
    for col,w in {"A":18,"B":34,"C":34,"D":34,"E":34,"F":36}.items():
        ws.column_dimensions[col].width=w

if __name__=="__main__":
    ap=argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--profile",choices=["reviewer","founder"],
                    help="build one profile (requires --out); omit to build all four")
    ap.add_argument("--lang",choices=["en","et"],default="en")
    ap.add_argument("--out",help="output .xlsx path (single-profile mode)")
    ap.add_argument("--outdir",default=".",
                    help="directory for the all-four build (default: current dir)")
    a=ap.parse_args()

    if a.profile:
        if not a.out:
            ap.error("--out is required with --profile")
        outdir=os.path.dirname(os.path.abspath(a.out))
        if not os.path.isdir(outdir):
            ap.error(f"output directory does not exist: {outdir}")
        print(f"{a.profile}/{a.lang}: {build(a.profile,a.lang,a.out)} items -> {a.out}")
    else:
        if not os.path.isdir(a.outdir):
            ap.error(f"output directory does not exist: {a.outdir}")
        for prof in ["reviewer","founder"]:
            for lang in ["en","et"]:
                fn=os.path.join(a.outdir,f"review_{prof}_{lang.upper()}.xlsx")
                print(f"{prof}/{lang}: {build(prof,lang,fn)} items -> {fn}")
